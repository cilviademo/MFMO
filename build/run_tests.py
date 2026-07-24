#!/usr/bin/env python3
"""Test harness for MFRP_v1.xlsx — runs the spec §12 test cases that apply to
Stage 1. Injects scenario data into a copy, recalculates with the `formulas`
engine, and asserts expected outputs. LibreOffice is blocked in this sandbox,
so TC-21 is covered by static prohibited-function analysis instead of a live
LibreOffice open."""
import os
import re
import sys
import tempfile
import openpyxl

sys.path.insert(0, "build")
from lib_calc import Calc  # noqa: E402

BASE = os.environ.get("MFRP_TEST_BOOK", "MFRP_v1.xlsx")
SCRATCH = "/tmp/claude-0/-home-user-MFMO/ce620f6d-3359-5659-b656-9d5284a081e9/scratchpad"
CT_FIRST = 8
CT_CLEAR_LAST = int(os.environ.get("MFRP_CT_LAST", "307"))
CW_CLEAR_LAST = int(os.environ.get("MFRP_CW_LAST", "200"))

results = []  # (id, ok, detail)


def num(v):
    try:
        return float(v)
    except Exception:
        return v


def approx(a, b, tol=1e-6):
    try:
        return abs(float(a) - float(b)) <= tol
    except Exception:
        return a == b


def make_book(cw_rows, ct_rows, filters=None):
    """Return path to a temp workbook with only the given crosswalk + tracker rows."""
    wb = openpyxl.load_workbook(BASE)
    cw = wb["Ref_Crosswalk"]
    for r in range(2, CW_CLEAR_LAST + 1):
        for col in "ABCDEFGHIJKLMNO":
            cw[f"{col}{r}"] = None
    for i, row in enumerate(cw_rows):
        r = 2 + i
        cw[f"A{r}"] = row["id"]
        cw[f"B{r}"] = row.get("name", row["id"])
        cw[f"C{r}"] = row.get("base", "BASE")
        cw[f"E{r}"] = row.get("majcom", "AETC")
        cw[f"F{r}"] = row["ftype"]
        cw[f"L{r}"] = "Y"
    ct = wb["Out_ComplianceTracker"]
    for r in range(CT_FIRST, CT_CLEAR_LAST + 1):
        for col in "ABCDEFGHI":
            ct[f"{col}{r}"] = None
    for i, row in enumerate(ct_rows):
        r = CT_FIRST + i
        ct[f"A{r}"] = row["dfac"]
        ct[f"B{r}"] = row["period"]
        ct[f"C{r}"] = row.get("ei")
        ct[f"D{r}"] = row.get("gl")
        ct[f"E{r}"] = row.get("cog")
        ct[f"F{r}"] = row.get("mfr")
        ct[f"G{r}"] = row.get("fss")
        ct[f"H{r}"] = row.get("inv")
    dash = wb["Out_Dashboard"]
    filters = filters or {}
    dash["C5"] = filters.get("maj")
    dash["C6"] = filters.get("base")
    dash["C7"] = filters.get("period")
    path = os.path.join(tempfile.mkdtemp(dir=SCRATCH), "t.xlsx")
    wb.save(path)
    return path


def ct(c, col, row):
    return c.get("Out_ComplianceTracker", f"{col}{row}")


def record(tid, ok, detail):
    results.append((tid, ok, detail))
    print(f"[{'PASS' if ok else 'FAIL'}] {tid}: {detail}")


# ---------------------------------------------------------------- TC-01
def tc01():
    p = make_book([{"id": "TEST-01", "ftype": "DFAC_LEGACY"}],
                  [{"dfac": "TEST-01", "period": "FY2026-P07A", "ei": 100000, "gl": 2500}])
    c = Calc(p)
    ok = (approx(ct(c, "K", 8), 0.03) and approx(ct(c, "M", 8), 0.025)
          and ct(c, "P", 8) == "PASS" and ct(c, "Q", 8) == "5.13")
    record("TC-01", ok, f"K={ct(c,'K',8)} M={ct(c,'M',8)} RESULT={ct(c,'P',8)} cite={ct(c,'Q',8)}")


# ---------------------------------------------------------------- TC-02
def tc02():
    cw = [{"id": "TEST-01", "ftype": "DFAC_LEGACY"}, {"id": "TEST-02", "ftype": "CAFE"}]
    # Part A: identical 2500 -> both PASS, different cites
    p = make_book(cw, [{"dfac": "TEST-01", "period": "FY2026-P07A", "ei": 100000, "gl": 2500},
                        {"dfac": "TEST-02", "period": "FY2026-P07A", "ei": 100000, "gl": 2500}])
    c = Calc(p)
    a_ok = (approx(ct(c, "K", 8), 0.03) and ct(c, "P", 8) == "PASS" and ct(c, "Q", 8) == "5.13"
            and approx(ct(c, "K", 9), 0.05) and ct(c, "P", 9) == "PASS" and ct(c, "Q", 9) == "6.6")
    record("TC-02a", a_ok, f"legacy(PASS,5.13)={ct(c,'P',8)},{ct(c,'Q',8)} "
           f"cafe(PASS,6.6)={ct(c,'P',9)},{ct(c,'Q',9)}")
    # Part B: identical 4000 -> legacy FAIL, cafe PASS (the critical divergence)
    p = make_book(cw, [{"dfac": "TEST-01", "period": "FY2026-P07A", "ei": 100000, "gl": 4000},
                        {"dfac": "TEST-02", "period": "FY2026-P07A", "ei": 100000, "gl": 4000}])
    c = Calc(p)
    b_ok = (ct(c, "P", 8) == "FAIL" and ct(c, "P", 9) == "PASS"
            and approx(ct(c, "M", 8), 0.04) and approx(ct(c, "M", 9), 0.04))
    record("TC-02b(critical)", b_ok, f"identical 4% -> legacy={ct(c,'P',8)} cafe={ct(c,'P',9)}")


# ---------------------------------------------------------------- TC-03
def tc03():
    p = make_book([{"id": "TEST-03", "ftype": "FOOD_2_0"}],
                  [{"dfac": "TEST-03", "period": "FY2026-P07A", "ei": 100000,
                    "cog": 44000, "gl": 8000}])
    c = Calc(p)
    ok = (ct(c, "L", 8) == "MARGIN" and approx(ct(c, "N", 8), 0.44)
          and ct(c, "P", 8) == "PASS")
    record("TC-03", ok, f"metric={ct(c,'L',8)} COG%={ct(c,'N',8)} RESULT={ct(c,'P',8)} "
           f"(8% gain/loss must not drive result)")


# ---------------------------------------------------------------- TC-04
def tc04():
    p = make_book([{"id": "TEST-04", "ftype": "ANG_DFAC"}],
                  [{"dfac": "TEST-04", "period": "FY2026-P07A", "ei": 50000, "gl": 4000}])
    c = Calc(p)
    ok = (approx(ct(c, "K", 8), 0.10) and approx(ct(c, "M", 8), 0.08)
          and ct(c, "P", 8) == "PASS")
    record("TC-04", ok, f"K={ct(c,'K',8)} M={ct(c,'M',8)} RESULT={ct(c,'P',8)}")


# ---------------------------------------------------------------- TC-05
def tc05():
    rows = [{"dfac": "TEST-01", "period": pk, "ei": 100000, "gl": 1000}
            for pk in ("FY2026-P01A", "FY2026-P01B", "FY2026-P02A")]
    p = make_book([{"id": "TEST-01", "ftype": "DFAC_LEGACY"}], rows)
    c = Calc(p)
    streaks = [num(ct(c, "O", 8 + i)) for i in range(3)]
    escs = [ct(c, "R", 8 + i) for i in range(3)]
    watch = [ct(c, "S", 8 + i) for i in range(3)]
    ok = (all(s == 0 for s in streaks) and all(e == "CLEAR" for e in escs)
          and all(w == "" for w in watch))
    record("TC-05", ok, f"streaks={streaks} esc={set(escs)} watch_empty={all(w=='' for w in watch)}")


# ---------------------------------------------------------------- TC-06
def tc06():
    seq = [("FY2026-P01A", 1000, "PASS", 0), ("FY2026-P01B", 5000, "FAIL", 1),
           ("FY2026-P02A", 6000, "FAIL", 2), ("FY2026-P02B", 7000, "FAIL", 3),
           ("FY2026-P03A", 1000, "PASS", 0), ("FY2026-P03B", 5000, "FAIL", 1)]
    rows = [{"dfac": "TEST-05", "period": pk, "ei": 100000, "gl": gl} for pk, gl, _, _ in seq]
    p = make_book([{"id": "TEST-05", "ftype": "DFAC_LEGACY"}], rows)
    c = Calc(p)
    exp_esc = {0: "CLEAR", 1: "MFR DUE - MANAGER", 2: "MFR DUE - FSS/CC SIGNATURE REQUIRED",
               3: "INVESTIGATION + REPORT OF SURVEY; MSG/CC NOTIFIED"}
    ok = True
    detail = []
    for i, (pk, gl, res, st) in enumerate(seq):
        r = 8 + i
        gotP, gotO, gotR = ct(c, "P", r), num(ct(c, "O", r)), ct(c, "R", r)
        line_ok = (gotP == res and gotO == st and gotR == exp_esc[st])
        ok = ok and line_ok
        detail.append(f"{pk}:{gotP}/streak{int(gotO)}")
    record("TC-06", ok, " ".join(detail) + " (reset@P03A verified)")


# ---------------------------------------------------------------- TC-07
def tc07():
    rows = [{"dfac": "TEST-05", "period": pk, "ei": 100000, "gl": gl}
            for pk, gl in [("FY2026-P01A", 5000), ("FY2026-P01B", 6000),
                           ("FY2026-P02A", 1000)]]
    # P01A FAIL(1), P01B FAIL(2 -> watch), P02A PASS(0)
    p = make_book([{"id": "TEST-05", "ftype": "DFAC_LEGACY"}], rows)
    c = Calc(p)
    watch_at_2 = ct(c, "S", 9)  # P01B streak 2
    watch_at_1 = ct(c, "S", 8)  # streak 1
    watch_at_0 = ct(c, "S", 10)  # streak 0
    a21 = c.get("Out_Dashboard", "A21")
    ok = ("WATCH" in (watch_at_2 or "") and watch_at_1 == "" and watch_at_0 == ""
          and a21 == "TEST-05")
    record("TC-07", ok, f"watchflag@streak2={bool(watch_at_2)} @1={bool(watch_at_1)} "
           f"@0={bool(watch_at_0)} dashboard_watchlist={a21!r}")


# ---------------------------------------------------------------- TC-08
def tc08():
    rows = [{"dfac": "TEST-05", "period": "FY2026-P01A", "ei": 100000, "gl": 5000,
             "mfr": "Y", "fss": "N"},
            {"dfac": "TEST-05", "period": "FY2026-P01B", "ei": 100000, "gl": 6000,
             "mfr": "Y", "fss": "N"}]
    p = make_book([{"id": "TEST-05", "ftype": "DFAC_LEGACY"}], rows)
    c = Calc(p)
    t_before = ct(c, "T", 9)  # streak 2, MFR=Y, FSS=N
    # now set FSS=Y
    rows[1]["fss"] = "Y"
    p2 = make_book([{"id": "TEST-05", "ftype": "DFAC_LEGACY"}], rows)
    c2 = Calc(p2)
    t_after = ct(c2, "T", 9)
    ok = (t_before == "FSS/CC SIGNATURE MISSING" and t_after == "")
    record("TC-08", ok, f"overdue(FSS=N)={t_before!r} -> overdue(FSS=Y)={t_after!r}")


# ---------------------------------------------------------------- TC-09
def tc09():
    p = make_book([{"id": "TEST-06", "ftype": "WIDGET_SHOP"}],
                  [{"dfac": "TEST-06", "period": "FY2026-P07A", "ei": 100000, "gl": 2500}])
    c = Calc(p)
    k = ct(c, "K", 8)
    rule6 = c.get("Out_ValidationRegister", "D9")  # rule 6 is row 9
    ok = ((k == "" or k is None) and ct(c, "P", 8) == "CHECK" and rule6 == "FAIL")
    record("TC-09", ok, f"K={k!r} RESULT={ct(c,'P',8)} rule6={rule6}")


# ---------------------------------------------------------------- TC-10
def tc10():
    p = make_book([{"id": "TEST-07", "ftype": "DFAC_LEGACY"}],
                  [{"dfac": "TEST-07", "period": "FY2026-P07A", "ei": 0, "gl": 500}])
    c = Calc(p)
    m = num(ct(c, "M", 8))
    errs = c.scan_errors()
    ok = (approx(m, 0) and len(errs) == 0)
    record("TC-10", ok, f"VAR_PCT={m} (denominator guarded) errors={len(errs)}")


# ---------------------------------------------------------------- TC-11
def tc11():
    p = make_book([{"id": "TEST-08", "ftype": "DFAC_LEGACY"}],
                  [{"dfac": "TEST-08", "period": "FY2026-P07A", "ei": 100000, "gl": None}])
    c = Calc(p)
    m = num(ct(c, "M", 8))
    ok = (approx(m, 0) and ct(c, "P", 8) == "PASS" and len(c.scan_errors()) == 0)
    record("TC-11", ok, f"blank GAIN_LOSS -> VAR_PCT={m} RESULT={ct(c,'P',8)} (N() coercion)")


# ---------------------------------------------------------------- TC-12 (Stage-1 analog)
def tc12():
    p = make_book([{"id": "REAL-01", "ftype": "DFAC_LEGACY"}],
                  [{"dfac": "GHOST-99", "period": "FY2026-P07A", "ei": 100000, "gl": 2500}])
    c = Calc(p)
    j = ct(c, "J", 8)
    rule1 = c.get("Out_ValidationRegister", "D4")  # rule 1 row 4
    ok = (j == "UNKNOWN" and ct(c, "P", 8) == "CHECK" and rule1 == "FAIL"
          and len(c.scan_errors()) == 0)
    record("TC-12(tracker analog)", ok,
           f"unknown DFAC -> FACILITY_TYPE={j!r} RESULT={ct(c,'P',8)} rule1={rule1} "
           f"(full Ledger_ProvenanceCheck is Stage 2)")


# ---------------------------------------------------------------- TC-19
def tc19():
    p = make_book([{"id": "TEST-01", "ftype": "DFAC_LEGACY"}],
                  [{"dfac": "TEST-01", "period": "FY2026-P07A", "ei": 100000, "gl": 1000}])
    c = Calc(p)
    d10 = num(c.get("Out_Dashboard", "D10"))
    a21 = c.get("Out_Dashboard", "A21")
    a41 = c.get("Out_Dashboard", "A41")
    ok = (d10 == 0 and "clear" in (a21 or "").lower() and "No overdue" in (a41 or ""))
    record("TC-19", ok, f"out_of_tol={d10} watch={a21!r} overdue={a41!r}")


# ---------------------------------------------------------------- TC-20
def tc20():
    c = Calc(BASE)
    # find calendar rows for FY2026-P07A and P07B
    cal = openpyxl.load_workbook(BASE)["Ref_Calendar"]
    idx = {}
    for r in range(2, 400):
        v = cal[f"A{r}"].value
        if v in ("FY2026-P07A", "FY2026-P07B"):
            idx[v] = r
    ra, rb = idx["FY2026-P07A"], idx["FY2026-P07B"]
    sa, ea = cal[f"E{ra}"].value, cal[f"F{ra}"].value
    sb, eb = cal[f"E{rb}"].value, cal[f"F{rb}"].value
    ok = (sa.day == 1 and ea.day == 15 and sb.day == 16 and eb.day == 31
          and sa.month == 7 and sb.month == 7)
    record("TC-20", ok, f"P07A={sa.date()}..{ea.date()}  P07B={sb.date()}..{eb.date()} "
           f"(14&15 Jul->P07A, 16&31 Jul->P07B)")


# ---------------------------------------------------------------- TC-21 (static)
PROHIBITED = ["XLOOKUP", "XMATCH", "SORT", "FILTER", "UNIQUE", "SEQUENCE",
              "TEXTAFTER", "TEXTBEFORE", "TEXTSPLIT"]


def tc21():
    wb = openpyxl.load_workbook(BASE)
    bad = []
    struct = []
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell_ in row:
                v = cell_.value
                if isinstance(v, str) and v.startswith("="):
                    up = v.upper()
                    for fn in PROHIBITED:
                        # word-boundary before "(" to avoid matching inside names
                        if re.search(r"(?<![A-Z0-9_\.])" + fn + r"\s*\(", up):
                            bad.append((ws.title, cell_.coordinate, fn))
                    if "[@" in v:
                        struct.append((ws.title, cell_.coordinate))
    ok = (not bad and not struct)
    record("TC-21(static)", ok,
           f"prohibited-fn hits={len(bad)} structured-ref hits={len(struct)} "
           f"(LibreOffice unavailable in sandbox; verified by static scan)")
    if bad:
        print("      offending:", bad[:10])


# ---------------------------------------------------------------- TC-22
def tc22():
    c = Calc(BASE)
    errs = c.scan_errors()
    record("TC-22", len(errs) == 0, f"naked error cells across recalced workbook={len(errs)}")
    if errs:
        print("      ", errs[:10])


# ---------------------------------------------------------------- TC-23
def tc23():
    p = make_book([{"id": "TEST-01", "ftype": "DFAC_LEGACY"}],
                  [{"dfac": "TEST-01", "period": "FY2026-P07A", "ei": 100000, "gl": 4000}])
    c1 = Calc(p)
    c2 = Calc(p)
    cells = [("Out_ComplianceTracker", x) for x in ("P8", "O8", "M8")] + \
            [("Out_Dashboard", x) for x in ("D10", "D15")]
    same = all(str(c1.get(s, co)) == str(c2.get(s, co)) for s, co in cells)
    record("TC-23(idempotence)", same,
           f"repeated recalc identical={same} (ledger TXN_ID de-dup is Stage 2)")


# ---------------------------------------------------------------- TC-24
def tc24():
    types = ["DFAC_LEGACY", "CAFE", "ANG_DFAC", "DFAC_LEGACY", "CAFE",
             "ANG_DFAC", "DFAC_LEGACY", "CAFE", "ANG_DFAC", "DFAC_LEGACY"]
    majcoms = ["AETC", "AETC", "AETC", "ACC", "ACC", "ACC", "AMC", "AMC", "AMC", "AETC"]
    cw = [{"id": f"ROLL-{i:02d}", "ftype": types[i], "majcom": majcoms[i],
           "base": majcoms[i] + "B"} for i in range(10)]
    # gain/loss 4000 on EI 100000 = 4%: FAIL for legacy(3%), PASS for cafe(5%)/ang(10%)
    ctr = [{"dfac": f"ROLL-{i:02d}", "period": "FY2026-P07A", "ei": 100000, "gl": 4000}
           for i in range(10)]
    # unfiltered
    p = make_book(cw, ctr)
    c = Calc(p)
    all_fail = num(c.get("Out_Dashboard", "D10"))  # legacy count = indices 0,3,6,9 => 4
    # filter to AETC (indices 0,1,2,9 -> types legacy,cafe,ang,legacy -> 2 legacy FAIL)
    p2 = make_book(cw, ctr, {"maj": "AETC"})
    c2 = Calc(p2)
    aetc_fail = num(c2.get("Out_Dashboard", "D10"))
    # verify per-facility tolerance: a cafe row PASSes at 4% while a legacy FAILs
    legacy_res = ct(c, "P", 8)   # ROLL-00 legacy
    cafe_res = ct(c, "P", 9)     # ROLL-01 cafe
    ok = (all_fail == 4 and aetc_fail == 2 and legacy_res == "FAIL" and cafe_res == "PASS")
    record("TC-24", ok, f"all_out_of_tol={all_fail} AETC_out_of_tol={aetc_fail} "
           f"legacy@4%={legacy_res} cafe@4%={cafe_res} (each scored on own tolerance)")


ALL = [tc01, tc02, tc03, tc04, tc05, tc06, tc07, tc08, tc09, tc10, tc11, tc12,
       tc19, tc20, tc21, tc22, tc23, tc24]

if __name__ == "__main__":
    for fn in ALL:
        try:
            fn()
        except Exception as e:
            record(fn.__name__, False, f"EXCEPTION {type(e).__name__}: {e}")
    npass = sum(1 for _, ok, _ in results if ok)
    print("\n" + "=" * 60)
    print(f"SUMMARY: {npass}/{len(results)} checks passed")
    print("NOT RUN (Stage 2 features, out of scope for this build): "
          "TC-13 provenance completeness, TC-14 adapter-version, TC-15..18 "
          "(ledger/reconciliation).")
    failed = [t for t, ok, _ in results if not ok]
    if failed:
        print("FAILED:", failed)
        sys.exit(1)
