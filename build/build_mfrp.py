#!/usr/bin/env python3
"""
Build tooling for MFRP_v1.xlsx — Mission Feeding Reconciliation Platform, Stage 1.

This script is BUILD TOOLING ONLY. The delivered artifact (MFRP_v1.xlsx) contains
formulas + data validation only, no macros, no Python dependency (spec §3).

Stage 1 tabs (spec §10):
  Ref_Tolerance, Ref_Calendar, Ref_Crosswalk, Ref_ExceptionTypes,
  Ref_AdapterManifest, Ref_CrosswalkVersion,
  Out_ComplianceTracker, Out_Dashboard, Out_DataDictionary, Out_ValidationRegister
"""

import calendar
import datetime
import os
import sys

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import CellIsRule, FormulaRule
from openpyxl.utils import get_column_letter

# -------------------------------------------------------------------------
# Build-date reference. The spec forbids hardcoded *results*; a static
# reference date for STATUS labelling in a hand-maintained calendar tab is data,
# not a computed result, and the user edits it freely.
# -------------------------------------------------------------------------
BUILD_DATE = datetime.date(2026, 7, 24)

# ---- Style constants (spec §9) ------------------------------------------
FONT_NAME = "Arial"

C_INPUT = "FF0000FF"     # blue text  = hardcoded input / user-editable
C_FORMULA = "FF000000"   # black text = formula
C_LINK = "FF008000"      # green text = cross-sheet link
FILL_YELLOW = "FFFFFF00"  # cells the user must fill in
FILL_RED = "FFFFC7CE"     # exception requiring action
FILL_AMBER = "FFFFEB9C"   # watch — approaching threshold
FILL_HDR = "FFD9D9D9"     # header band grey
FILL_TITLE = "FF1F3864"   # dark blue title band
FILL_SEED = "FFF2F2F2"    # seeded reference data (read-only) light grey

TAB_GREY = "808080"       # Reference tabs
TAB_DARKBLUE = "1F3864"   # Output tabs

FMT_CCY = "$#,##0;($#,##0);-"
FMT_PCT = "0.0%"
FMT_DATE = "DD-MMM-YYYY"
FMT_INT = "#,##0"
FMT_YEAR = "0"            # years as text-like integer, never 2,026

MAJCOMS = "AETC,ACC,AMC,PACAF,USAFE,AFGSC,AFMC,AFSOC,ANG,AFRC"

thin = Side(style="thin", color="FFB0B0B0")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)


def font(bold=False, color=C_FORMULA, size=10, italic=False):
    return Font(name=FONT_NAME, bold=bold, color=color, size=size, italic=italic)


def cell(ws, coord, value=None, f=None, fill=None, fmt=None, align=None,
         wrap=False, border=False):
    c = ws[coord]
    if value is not None:
        c.value = value
    c.font = f if f else font()
    if fill:
        c.fill = PatternFill("solid", fgColor=fill)
    if fmt:
        c.number_format = fmt
    if align or wrap:
        c.alignment = Alignment(horizontal=align, vertical="top", wrap_text=wrap)
    if border:
        c.border = BORDER
    return c


def header_row(ws, row, headers, start_col=1):
    for i, h in enumerate(headers):
        col = get_column_letter(start_col + i)
        cell(ws, f"{col}{row}", h, f=font(bold=True, color="FFFFFFFF"),
             fill=FILL_TITLE, wrap=True, border=True, align="center")


def title(ws, text, sub=None):
    cell(ws, "A1", text, f=font(bold=True, color="FFFFFFFF", size=14),
         fill=FILL_TITLE)
    if sub:
        cell(ws, "A2", sub, f=font(italic=True, color="FF404040", size=9))


# =========================================================================
wb = openpyxl.Workbook()
wb.remove(wb.active)


# -------------------------------------------------------------------------
# 1. Ref_Tolerance   (columns: A FACILITY_TYPE, B TOLERANCE_PCT, C BASIS,
#    D METRIC_TYPE, E DAFMAN_CITE, F TIER, G NOTES) — anchors the §8.1 formulas
# -------------------------------------------------------------------------
CONFLICT_NOTE = ("Para 5.13 sets 3% for DFAC food accounts; para 6.6 sets 5% for "
                 "CAFE facilities. Both T-2. Applied standard depends on facility "
                 "classification. Inconsistency surfaced intentionally.")
FOOD20_NOTE = ("45% COG is a pricing/margin standard, not a variance tolerance. "
               "Food 2.0 sites have no AvT-equivalent variance control. A site can "
               "meet 45% COG exactly and still carry large theoretical-vs-actual "
               "variance.")

TOL_ROWS = [
    ("DFAC_LEGACY", 0.03, "MONTHLY_EARNED_INCOME", "VARIANCE", "5.13", "T-2", CONFLICT_NOTE),
    ("CAFE", 0.05, "TOTAL_EARNED_INCOME", "VARIANCE", "6.6", "T-2", CONFLICT_NOTE),
    ("FOOD_2_0", 0.45, "COST_OF_GOODS", "MARGIN", "6.6, 3.13.3", "T-2", FOOD20_NOTE),
    ("ANG_DFAC", 0.10, "TOTAL_EARNED_INCOME", "VARIANCE", "6.6", "T-2", ""),
    ("FIELD_FEEDING", 0.03, "TOTAL_EARNED_INCOME", "VARIANCE", "5.18.6", "T-2", ""),
    ("NGB_UTA", 0.10, "MONTHLY_EARNED_INCOME", "VARIANCE", "5.13.1", "T-2", ""),
    ("NGB_AT", 0.05, "MONTHLY_EARNED_INCOME", "VARIANCE", "5.13.1", "T-2", ""),
    ("NGB_EXTENDED", 0.03, "MONTHLY_EARNED_INCOME", "VARIANCE", "5.13.1", "T-2", ""),
]


def build_ref_tolerance():
    ws = wb.create_sheet("Ref_Tolerance")
    ws.sheet_properties.tabColor = TAB_GREY
    heads = ["FACILITY_TYPE", "TOLERANCE_PCT", "BASIS", "METRIC_TYPE",
             "DAFMAN_CITE", "TIER", "NOTES"]
    header_row(ws, 1, heads)
    for i, (ft, tol, basis, metric, cite, tier, note) in enumerate(TOL_ROWS):
        r = 2 + i
        cell(ws, f"A{r}", ft, f=font(bold=True), fill=FILL_SEED, border=True)
        cell(ws, f"B{r}", tol, fmt=FMT_PCT, fill=FILL_SEED, border=True,
             align="center")
        cell(ws, f"C{r}", basis, fill=FILL_SEED, border=True)
        cell(ws, f"D{r}", metric, f=font(bold=(metric == "MARGIN")),
             fill=FILL_SEED, border=True, align="center")
        cell(ws, f"E{r}", cite, fill=FILL_SEED, border=True, align="center")
        cell(ws, f"F{r}", tier, fill=FILL_SEED, border=True, align="center")
        cell(ws, f"G{r}", note, fill=FILL_SEED, border=True, wrap=True)
    # Legend / notice
    r = 2 + len(TOL_ROWS) + 1
    cell(ws, f"A{r}", "REGULATORY TABLE — DO NOT EDIT TOLERANCE VALUES.",
         f=font(bold=True, color="FFC00000"))
    cell(ws, f"A{r+1}",
         "Tolerances are set by DAFMAN 34-131. To change how a facility is scored, "
         "change its FACILITY_TYPE in Ref_Crosswalk, not a value here.",
         f=font(italic=True, size=9))
    cell(ws, f"A{r+2}",
         "TOLERANCE_PCT is stored as a fraction (0.03 renders 3.0%). "
         "METRIC_TYPE MARGIN (Food 2.0) is tested on cost-of-goods, not variance.",
         f=font(italic=True, size=9))
    widths = {"A": 16, "B": 14, "C": 24, "D": 13, "E": 14, "F": 7, "G": 70}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w
    ws.freeze_panes = "A2"


# -------------------------------------------------------------------------
# 2. Ref_Calendar  (A PERIOD_KEY, B FY, C MONTH, D HALF, E START_DATE,
#    F END_DATE, G INVENTORY_DUE, H POST_DUE, I MMR_DUE, J STATUS)
# -------------------------------------------------------------------------
def build_ref_calendar():
    ws = wb.create_sheet("Ref_Calendar")
    ws.sheet_properties.tabColor = TAB_GREY
    heads = ["PERIOD_KEY", "FY", "MONTH", "HALF", "START_DATE", "END_DATE",
             "INVENTORY_DUE", "POST_DUE", "MMR_DUE", "STATUS"]
    header_row(ws, 1, heads)
    r = 2
    # calendar months Oct 2024 (FY2025 start) .. Sep 2028 (FY2028 end)
    ym = []
    y, m = 2024, 10
    while (y, m) <= (2028, 9):
        ym.append((y, m))
        m += 1
        if m == 13:
            m = 1
            y += 1
    for (yy, mm) in ym:
        fy = yy + 1 if mm >= 10 else yy
        last = calendar.monthrange(yy, mm)[1]
        nm_y, nm_m = (yy + 1, 1) if mm == 12 else (yy, mm + 1)
        mmr_due = datetime.date(nm_y, nm_m, 10)
        for half in ("A", "B"):
            if half == "A":
                start, end = datetime.date(yy, mm, 1), datetime.date(yy, mm, 15)
            else:
                start, end = datetime.date(yy, mm, 16), datetime.date(yy, mm, last)
            inv_due = end
            post_due = end + datetime.timedelta(days=5)
            status = "CLOSED" if end < BUILD_DATE else "OPEN"
            pk = f"FY{fy}-P{mm:02d}{half}"
            cell(ws, f"A{r}", pk, f=font(bold=True), fill=FILL_SEED, border=True)
            cell(ws, f"B{r}", fy, fmt=FMT_YEAR, fill=FILL_SEED, border=True, align="center")
            cell(ws, f"C{r}", mm, fmt="0", fill=FILL_SEED, border=True, align="center")
            cell(ws, f"D{r}", half, fill=FILL_SEED, border=True, align="center")
            cell(ws, f"E{r}", start, fmt=FMT_DATE, fill=FILL_SEED, border=True)
            cell(ws, f"F{r}", end, fmt=FMT_DATE, fill=FILL_SEED, border=True)
            cell(ws, f"G{r}", inv_due, fmt=FMT_DATE, fill=FILL_SEED, border=True)
            cell(ws, f"H{r}", post_due, fmt=FMT_DATE, fill=FILL_SEED, border=True)
            cell(ws, f"I{r}", mmr_due, fmt=FMT_DATE, fill=FILL_SEED, border=True)
            cell(ws, f"J{r}", status, fill=FILL_SEED, border=True, align="center")
            r += 1
    cell(ws, f"A{r+1}",
         "Semi-monthly periods (para 7.13 inventory on 15th & last day; para 5.2.5 "
         "MMR due 10th of following month). STATUS is maintained by the build owner.",
         f=font(italic=True, size=9))
    widths = {"A": 14, "B": 7, "C": 7, "D": 6, "E": 13, "F": 13, "G": 14,
              "H": 13, "I": 13, "J": 9}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w
    ws.freeze_panes = "A2"


# -------------------------------------------------------------------------
# 3. Ref_Crosswalk  (A DFAC_ID .. O NOTES) — anchors §8.1 lookups (A, C, E, F)
# -------------------------------------------------------------------------
CW_HEADS = ["DFAC_ID", "DFAC_NAME", "BASE_ID", "BASE_NAME", "MAJCOM",
            "FACILITY_TYPE", "DODAAC", "ALOHA_STORE", "CT_SITE", "CIR_MERCHANT",
            "STORES_ACCT", "ACTIVE", "EFF_START", "EFF_END", "NOTES"]
CW_PLACEHOLDERS = {"H", "I", "J", "K"}  # ALOHA_STORE, CT_SITE, CIR_MERCHANT, STORES_ACCT
CW_LAST_ENTRY_ROW = int(os.environ.get("MFRP_CW_LAST", "200"))


def build_ref_crosswalk():
    ws = wb.create_sheet("Ref_Crosswalk")
    ws.sheet_properties.tabColor = TAB_GREY
    # legend row 1 handled via header on row 2; put a legend line above? Keep header at row1
    header_row(ws, 1, CW_HEADS)
    # Example row (row 2) — realistic values, blue input font
    ex = ["JBSA-01", "Gateway Inn", "JBSA", "Joint Base San Antonio", "AETC",
          "DFAC_LEGACY", "FA1234", "", "", "", "", "Y",
          datetime.date(2025, 10, 1), None, "Example row — replace with real DFACs"]
    for i, val in enumerate(ex):
        col = get_column_letter(1 + i)
        is_ph = col in CW_PLACEHOLDERS
        f = font(color=C_INPUT)
        fmt = FMT_DATE if col in ("M", "N") else None
        cell(ws, f"{col}2", val if val is not None else None, f=f,
             fill=(FILL_YELLOW if is_ph else None), fmt=fmt, border=True)
    cell(ws, "P2", "◄ EXAMPLE", f=font(italic=True, color="FFC00000", size=9))
    # Blank yellow entry rows 3..CW_LAST_ENTRY_ROW
    for r in range(3, CW_LAST_ENTRY_ROW + 1):
        for i, col in enumerate([get_column_letter(1 + j) for j in range(len(CW_HEADS))]):
            fmt = FMT_DATE if col in ("M", "N") else None
            cell(ws, f"{col}{r}", None, f=font(color=C_INPUT),
                 fill=FILL_YELLOW, fmt=fmt, border=True)
    # Data validation
    dv_ft = DataValidation(type="list", formula1="'Ref_Tolerance'!$A$2:$A$9",
                           allow_blank=True)
    dv_ft.error = "Must be a FACILITY_TYPE defined in Ref_Tolerance."
    dv_ft.prompt = "Pick a facility type (drives which tolerance applies)."
    ws.add_data_validation(dv_ft)
    dv_ft.add(f"F2:F{CW_LAST_ENTRY_ROW}")
    dv_maj = DataValidation(type="list", formula1=f'"{MAJCOMS}"', allow_blank=True)
    ws.add_data_validation(dv_maj)
    dv_maj.add(f"E2:E{CW_LAST_ENTRY_ROW}")
    dv_yn = DataValidation(type="list", formula1='"Y,N"', allow_blank=True)
    ws.add_data_validation(dv_yn)
    dv_yn.add(f"L2:L{CW_LAST_ENTRY_ROW}")
    # Legend below the entry area
    lr = CW_LAST_ENTRY_ROW + 2
    cell(ws, f"A{lr}", "LEGEND", f=font(bold=True))
    cell(ws, f"A{lr+1}", "Yellow cells = fill these in. Placeholder columns "
         "(ALOHA_STORE, CT_SITE, CIR_MERCHANT, STORES_ACCT) await real system "
         "identifiers — leave blank until you have real exports.",
         f=font(italic=True, size=9))
    cell(ws, f"A{lr+2}", "DFAC_ID is the canonical key, format {BASE_ID}-{NN} "
         "(e.g. JBSA-01). Once assigned, never change one — retire via EFF_END and "
         "reissue. FACILITY_TYPE must match Ref_Tolerance exactly.",
         f=font(italic=True, size=9))
    widths = {"A": 12, "B": 18, "C": 10, "D": 22, "E": 9, "F": 14, "G": 10,
              "H": 12, "I": 12, "J": 13, "K": 12, "L": 8, "M": 12, "N": 12, "O": 30}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w
    ws.freeze_panes = "A2"


# -------------------------------------------------------------------------
# 4. Ref_ExceptionTypes (seed reference — §5.6)
# -------------------------------------------------------------------------
EXC_ROWS = [
    ("LINK_ERR", "Item linked incorrectly", "CT_VS_ALOHA", "AFSVC/VMF",
     "Route to VMF — bases cannot self-correct links (para 7.8.1)"),
    ("PRICE_ERR", "Incorrect POS price", "CT_VS_ALOHA", "AFSVC/VMF",
     "Pricing centrally managed (para 3.13.1, 6.4)"),
    ("VOID_ERR", "Void not propagated", "CT_VS_ALOHA", "Base accountant",
     "Verify POS void handling"),
    ("ADJ_ERR", "Wrong inventory adjustment", "CT_VS_ALOHA", "Base DFAC manager",
     "Review raw/finished waste entries (para 7.10)"),
    ("PURCH_VAR", "Purchase vs SF 1080 mismatch", "PURCHASES", "Base accountant",
     "Purchase Reconciliation Worksheet (para 7.18)"),
    ("MISSING_SRC", "Expected source file absent", "INGEST", "Base accountant",
     "Check export/upload"),
    ("PROV_NULL", "Provenance incomplete", "INTEGRITY", "Build owner",
     "Adapter defect — fix immediately"),
    ("XWALK_UNRESOLVED", "DFAC_ID not in crosswalk", "INTEGRITY", "AFSVC/VMF",
     "Add to crosswalk registry"),
    ("TOL_BREACH", "Out of tolerance", "COMPLIANCE", "FSO/FSSC",
     "See escalation ladder (para 5.13.2)"),
]


def build_ref_exceptiontypes():
    ws = wb.create_sheet("Ref_ExceptionTypes")
    ws.sheet_properties.tabColor = TAB_GREY
    heads = ["EXC_CODE", "EXC_NAME", "CATEGORY", "DEFAULT_OWNER", "REMEDIATION"]
    header_row(ws, 1, heads)
    for i, row in enumerate(EXC_ROWS):
        r = 2 + i
        for j, val in enumerate(row):
            col = get_column_letter(1 + j)
            cell(ws, f"{col}{r}", val, f=font(bold=(j == 0)), fill=FILL_SEED,
                 border=True, wrap=(j == 4))
    r = 2 + len(EXC_ROWS) + 1
    cell(ws, f"A{r}", "Seed taxonomy (spec §5.6). Referenced by the reconciliation "
         "layer in Stage 3; ownership routing enforced there.",
         f=font(italic=True, size=9))
    widths = {"A": 18, "B": 26, "C": 14, "D": 18, "E": 55}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w
    ws.freeze_panes = "A2"


# -------------------------------------------------------------------------
# 5. Ref_AdapterManifest (hand-maintained edge list — §5.4)
# -------------------------------------------------------------------------
def build_ref_adaptermanifest():
    ws = wb.create_sheet("Ref_AdapterManifest")
    ws.sheet_properties.tabColor = TAB_GREY
    heads = ["QUERY_NAME", "LAYER", "SOURCE_SYSTEM", "INPUT_PATH", "OUTPUT_TABLE",
             "ADAPTER_VER", "OWNER", "LAST_CHANGED", "STATUS", "TEST_FILE", "NOTES"]
    header_row(ws, 1, heads)
    # Example STUB row
    ex = ["adapter_DFAS_SIK_v2", "ADAPTER", "DFAS", "\\ingest\\dfas\\",
          "ledger_Transactions", "DFAS_v2", "AFSVC/VMF",
          datetime.date(2026, 8, 1), "STUB", "\\tests\\dfas_sample.pdf",
          "Example — adapters remain stubs until real headers arrive (§14)"]
    for i, val in enumerate(ex):
        col = get_column_letter(1 + i)
        fmt = FMT_DATE if col == "H" else None
        cell(ws, f"{col}2", val, f=font(color=C_INPUT),
             fmt=fmt, border=True, fill=FILL_YELLOW)
    cell(ws, "M2", "◄ EXAMPLE", f=font(italic=True, color="FFC00000", size=9))
    for r in range(3, 40):
        for i in range(len(heads)):
            col = get_column_letter(1 + i)
            fmt = FMT_DATE if col == "H" else None
            cell(ws, f"{col}{r}", None, f=font(color=C_INPUT), fill=FILL_YELLOW,
                 fmt=fmt, border=True)
    dv_layer = DataValidation(type="list",
                              formula1='"ADAPTER,LEDGER,RECON,OUTPUT"', allow_blank=True)
    ws.add_data_validation(dv_layer)
    dv_layer.add("B2:B39")
    dv_status = DataValidation(type="list",
                               formula1='"ACTIVE,DEPRECATED,STUB"', allow_blank=True)
    ws.add_data_validation(dv_status)
    dv_status.add("I2:I39")
    cell(ws, "A41", "Hand-maintained transformation graph. Register each Power Query "
         "here with a path to a known-good TEST_FILE. Adapters are Stage 2+; none are "
         "ACTIVE yet.", f=font(italic=True, size=9))
    widths = {"A": 22, "B": 10, "C": 14, "D": 16, "E": 18, "F": 12, "G": 12,
              "H": 13, "I": 11, "J": 20, "K": 40}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w
    ws.freeze_panes = "A2"


# -------------------------------------------------------------------------
# 6. Ref_CrosswalkVersion (§5.5)
# -------------------------------------------------------------------------
def build_ref_crosswalkversion():
    ws = wb.create_sheet("Ref_CrosswalkVersion")
    ws.sheet_properties.tabColor = TAB_GREY
    heads = ["CROSSWALK_VER", "EFFECTIVE_DATE", "CHANGED_BY", "CHANGE_SUMMARY",
             "ROW_COUNT"]
    header_row(ws, 1, heads)
    ex = ["CW-2026-08-01", datetime.date(2026, 8, 1), "AFSVC/VMF",
          "Initial registry", 1]
    for i, val in enumerate(ex):
        col = get_column_letter(1 + i)
        fmt = FMT_DATE if col == "B" else (FMT_INT if col == "E" else None)
        cell(ws, f"{col}2", val, f=font(color=C_INPUT), fmt=fmt,
             fill=FILL_YELLOW, border=True)
    cell(ws, "G2", "◄ EXAMPLE", f=font(italic=True, color="FFC00000", size=9))
    for r in range(3, 30):
        for i in range(len(heads)):
            col = get_column_letter(1 + i)
            fmt = FMT_DATE if col == "B" else (FMT_INT if col == "E" else None)
            cell(ws, f"{col}{r}", None, f=font(color=C_INPUT), fill=FILL_YELLOW,
                 fmt=fmt, border=True)
    cell(ws, "A31", "Every crosswalk change gets a new version row. Ledger rows stamp "
         "the CROSSWALK_VER that resolved them, so silent mapping drift is detectable "
         "(Stage 2+).", f=font(italic=True, size=9))
    widths = {"A": 16, "B": 15, "C": 14, "D": 40, "E": 11}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w
    ws.freeze_panes = "A2"


# -------------------------------------------------------------------------
# 7. Out_ComplianceTracker  — the core engine (§8.1)
# Input:  A DFAC_ID  B PERIOD_KEY  C EARNED_INCOME  D GAIN_LOSS_AMT
#         E COST_OF_GOODS  F MFR_SUBMITTED  G FSS_CC_SIGNED  H INVESTIGATION_OPENED
#         I NOTES
# Calc:   J FACILITY_TYPE  K TOLERANCE_PCT  L METRIC_TYPE  M VAR_PCT  N COG_PCT
#         O STREAK  P RESULT  Q DAFMAN_CITE  R ESCALATION  S WATCH  T ACTION_OVERDUE
#         U SEQ(helper) V LASTPASS(helper) W MAJCOM(helper) X BASE_ID(helper)
#         Y MATCHF(helper) Z WATCH_RANK(helper) AA OVERDUE_RANK(helper)
# Data rows 8..CT_LAST (row 8 = first data row per spec).
# -------------------------------------------------------------------------
CT_FIRST = 8
CT_LAST = int(os.environ.get("MFRP_CT_LAST", "307"))
CT_HEADS = ["DFAC_ID", "PERIOD_KEY", "EARNED_INCOME", "GAIN_LOSS_AMT",
            "COST_OF_GOODS", "MFR_SUBMITTED", "FSS_CC_SIGNED",
            "INVESTIGATION_OPENED", "NOTES", "FACILITY_TYPE", "TOLERANCE_PCT",
            "METRIC_TYPE", "VAR_PCT", "COG_PCT", "STREAK", "RESULT", "DAFMAN_CITE",
            "ESCALATION", "WATCH", "ACTION_OVERDUE", "SEQ (helper)",
            "LASTPASS (helper)", "MAJCOM (helper)", "BASE_ID (helper)",
            "MATCHF (helper)", "WATCH_RANK (helper)", "OVERDUE_RANK (helper)"]


def ct_formulas(r):
    """Return dict col-letter -> formula string for data row r."""
    A = f"$A{r}"
    rng = lambda col: f"${col}${CT_FIRST}:${col}${CT_LAST}"
    f = {}
    f["J"] = (f'=IF({A}="","",IFERROR(INDEX(\'Ref_Crosswalk\'!$F$2:$F$500,'
              f'MATCH({A},\'Ref_Crosswalk\'!$A$2:$A$500,0)),"UNKNOWN"))')
    f["K"] = (f'=IF({A}="","",IFERROR(INDEX(\'Ref_Tolerance\'!$B$2:$B$9,'
              f'MATCH($J{r},\'Ref_Tolerance\'!$A$2:$A$9,0)),""))')
    f["L"] = (f'=IF({A}="","",IFERROR(INDEX(\'Ref_Tolerance\'!$D$2:$D$9,'
              f'MATCH($J{r},\'Ref_Tolerance\'!$A$2:$A$9,0)),""))')
    f["M"] = f'=IF({A}="","",IFERROR(ABS(N($D{r}))/N($C{r}),0))'
    f["N"] = f'=IF({A}="","",IFERROR(N($E{r})/N($C{r}),0))'
    f["O"] = (f'=IF({A}="",0,IFERROR(SUMPRODUCT(({rng("A")}={A})*'
              f'({rng("P")}="FAIL")*({rng("U")}>$V{r})*({rng("U")}<=$U{r})),0))')
    f["P"] = (f'=IF({A}="","",IFERROR(IF(NOT(ISNUMBER($K{r})),"CHECK",'
              f'IF($L{r}="MARGIN",IF($N{r}<=$K{r},"PASS","FAIL"),'
              f'IF($M{r}<=$K{r},"PASS","FAIL"))),"CHECK"))')
    f["Q"] = (f'=IF({A}="","",IFERROR(INDEX(\'Ref_Tolerance\'!$E$2:$E$9,'
              f'MATCH($J{r},\'Ref_Tolerance\'!$A$2:$A$9,0)),""))')
    f["R"] = (f'=IF({A}="","",IFERROR(IF($O{r}=0,"CLEAR",'
              f'IF($O{r}=1,"MFR DUE - MANAGER",'
              f'IF($O{r}=2,"MFR DUE - FSS/CC SIGNATURE REQUIRED",'
              f'"INVESTIGATION + REPORT OF SURVEY; MSG/CC NOTIFIED"))),""))')
    f["S"] = (f'=IF({A}="","",IF($O{r}=2,'
              f'"WATCH - ONE MORE MISS TRIGGERS INVESTIGATION",""))')
    f["T"] = (f'=IF({A}="","",IFERROR(IF(AND($O{r}>=1,$F{r}<>"Y"),'
              f'"MFR NOT SUBMITTED",IF(AND($O{r}>=2,$G{r}<>"Y"),'
              f'"FSS/CC SIGNATURE MISSING",IF(AND($O{r}>=3,$H{r}<>"Y"),'
              f'"INVESTIGATION NOT OPENED",""))),""))')
    f["U"] = (f'=IF({A}="",0,IFERROR(MATCH($B{r},'
              f'\'Ref_Calendar\'!$A$2:$A$500,0),0))')
    f["V"] = (f'=IF({A}="",0,IFERROR(SUMPRODUCT(MAX(({rng("A")}={A})*'
              f'({rng("P")}="PASS")*({rng("U")}<=$U{r})*{rng("U")})),0))')
    f["W"] = (f'=IF({A}="","",IFERROR(INDEX(\'Ref_Crosswalk\'!$E$2:$E$500,'
              f'MATCH({A},\'Ref_Crosswalk\'!$A$2:$A$500,0)),""))')
    f["X"] = (f'=IF({A}="","",IFERROR(INDEX(\'Ref_Crosswalk\'!$C$2:$C$500,'
              f'MATCH({A},\'Ref_Crosswalk\'!$A$2:$A$500,0)),""))')
    f["Y"] = (f'=IF({A}="",0,IF(AND('
              f'OR(\'Out_Dashboard\'!$C$5="",$W{r}=\'Out_Dashboard\'!$C$5),'
              f'OR(\'Out_Dashboard\'!$C$6="",$X{r}=\'Out_Dashboard\'!$C$6),'
              f'OR(\'Out_Dashboard\'!$C$7="",$B{r}=\'Out_Dashboard\'!$C$7)),1,0))')
    # incremental ranks use growing ranges from CT_FIRST to r
    f["Z"] = (f'=IF(AND({A}<>"",$S{r}<>"",$Y{r}=1),'
              f'SUMPRODUCT(($A${CT_FIRST}:$A{r}<>"")*($S${CT_FIRST}:$S{r}<>"")*'
              f'($Y${CT_FIRST}:$Y{r}=1)),"")')
    f["AA"] = (f'=IF(AND({A}<>"",$T{r}<>"",$Y{r}=1),'
               f'SUMPRODUCT(($A${CT_FIRST}:$A{r}<>"")*($T${CT_FIRST}:$T{r}<>"")*'
               f'($Y${CT_FIRST}:$Y{r}=1)),"")')
    return f


def build_out_compliancetracker():
    ws = wb.create_sheet("Out_ComplianceTracker")
    ws.sheet_properties.tabColor = TAB_DARKBLUE
    title(ws, "MISSION FEEDING — COMPLIANCE TRACKER",
          "One row per DFAC per period. Enter the yellow input columns; everything "
          "right of NOTES is computed. Then press Data > Refresh (F9).")
    # legend rows 3-6
    cell(ws, "A3", "HOW TO USE:", f=font(bold=True))
    cell(ws, "A4", "1) Pick DFAC_ID and PERIOD_KEY (dropdowns).  2) Enter "
         "EARNED_INCOME and GAIN_LOSS_AMT from the Monthly Monetary Record.  "
         "3) COST_OF_GOODS for Food 2.0 sites only.  4) Mark MFR/FSS_CC/INVESTIGATION "
         "as those actions occur.", f=font(size=9))
    cell(ws, "A5", "Blue = you type it (yellow cells).   Black = computed, do not edit.  "
         "Red fill = act now.   Amber fill = watch.", f=font(size=9))
    cell(ws, "A6", "Tolerance selection, pass/fail, streak counting and escalation are "
         "all computed. Do not hand-calculate any of it.", f=font(italic=True, size=9))
    # header row 7
    header_row(ws, 7, CT_HEADS)
    input_cols = ["A", "B", "C", "D", "E", "F", "G", "H", "I"]
    ccy_cols = {"C", "D", "E"}
    pct_cols = {"K", "M", "N"}
    helper_cols = {"U", "V", "W", "X", "Y", "Z", "AA"}
    # Example row 8
    example = {"A": "JBSA-01", "B": "FY2026-P07A", "C": 100000, "D": 2500,
               "E": None, "F": "Y", "G": "Y", "H": "N",
               "I": "Example row — clean PASS (2.5% <= 3.0%)"}
    for r in range(CT_FIRST, CT_LAST + 1):
        # input columns
        for col in input_cols:
            val = example.get(col) if r == CT_FIRST else None
            fmt = FMT_CCY if col in ccy_cols else None
            cell(ws, f"{col}{r}", val, f=font(color=C_INPUT),
                 fill=(None if r == CT_FIRST else FILL_YELLOW), fmt=fmt, border=True)
        # calc columns
        fdict = ct_formulas(r)
        for col, formula in fdict.items():
            fmt = FMT_PCT if col in pct_cols else None
            fcolor = "FF808080" if col in helper_cols else C_FORMULA
            cell(ws, f"{col}{r}", formula, f=font(color=fcolor), fmt=fmt, border=True)
    cell(ws, f"P{CT_FIRST}", ct_formulas(CT_FIRST)["P"])  # ensure example computed
    # note on example row
    cell(ws, f"AB{CT_FIRST}", "◄ EXAMPLE",
         f=font(italic=True, color="FFC00000", size=9))
    # Data validations
    dv_dfac = DataValidation(type="list", formula1="'Ref_Crosswalk'!$A$2:$A$500",
                             allow_blank=True)
    ws.add_data_validation(dv_dfac)
    dv_dfac.add(f"A{CT_FIRST}:A{CT_LAST}")
    dv_pk = DataValidation(type="list", formula1="'Ref_Calendar'!$A$2:$A$500",
                           allow_blank=True)
    ws.add_data_validation(dv_pk)
    dv_pk.add(f"B{CT_FIRST}:B{CT_LAST}")
    dv_yn = DataValidation(type="list", formula1='"Y,N"', allow_blank=True)
    ws.add_data_validation(dv_yn)
    dv_yn.add(f"F{CT_FIRST}:H{CT_LAST}")
    # Conditional formatting: action=red, watch=amber
    red = PatternFill("solid", fgColor=FILL_RED)
    amber = PatternFill("solid", fgColor=FILL_AMBER)
    ws.conditional_formatting.add(
        f"P{CT_FIRST}:P{CT_LAST}",
        CellIsRule(operator="equal", formula=['"FAIL"'], fill=red))
    ws.conditional_formatting.add(
        f"S{CT_FIRST}:S{CT_LAST}",
        FormulaRule(formula=[f'LEN($S{CT_FIRST})>0'], fill=amber))
    ws.conditional_formatting.add(
        f"T{CT_FIRST}:T{CT_LAST}",
        FormulaRule(formula=[f'LEN($T{CT_FIRST})>0'], fill=red))
    widths = {"A": 12, "B": 13, "C": 13, "D": 13, "E": 13, "F": 9, "G": 9, "H": 11,
              "I": 24, "J": 14, "K": 11, "L": 11, "M": 10, "N": 10, "O": 8, "P": 8,
              "Q": 11, "R": 34, "S": 30, "T": 22, "U": 10, "V": 12, "W": 10, "X": 10,
              "Y": 10, "Z": 12, "AA": 13}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w
    ws.freeze_panes = "C8"


# -------------------------------------------------------------------------
# 8. Out_Dashboard — the "What's Broken" screen (§8.2)
# -------------------------------------------------------------------------
def build_out_dashboard():
    ws = wb.create_sheet("Out_Dashboard")
    ws.sheet_properties.tabColor = TAB_DARKBLUE
    CT = "'Out_ComplianceTracker'!"
    rP = f"{CT}$P$8:$P$307"
    rO = f"{CT}$O$8:$O$307"
    rY = f"{CT}$Y$8:$Y$307"
    rS = f"{CT}$S$8:$S$307"
    rT = f"{CT}$T$8:$T$307"
    rZ = f"{CT}$Z$8:$Z$307"
    rAA = f"{CT}$AA$8:$AA$307"
    rA = f"{CT}$A$8:$A$307"

    cell(ws, "A1", "MISSION FEEDING — STATUS DASHBOARD",
         f=font(bold=True, color="FFFFFFFF", size=16), fill=FILL_TITLE)
    cell(ws, "A2", "Status as of:", f=font(bold=True))
    cell(ws, "C2", "=TODAY()", f=font(bold=True), fmt="DD-MMM-YYYY")

    cell(ws, "A4", "FILTERS  (leave blank = all)", f=font(bold=True))
    cell(ws, "A5", "MAJCOM:", f=font())
    cell(ws, "C5", None, f=font(color=C_INPUT), fill=FILL_YELLOW, border=True)
    cell(ws, "A6", "BASE_ID:", f=font())
    cell(ws, "C6", None, f=font(color=C_INPUT), fill=FILL_YELLOW, border=True)
    cell(ws, "A7", "PERIOD_KEY:", f=font())
    cell(ws, "C7", "FY2026-P07A", f=font(color=C_INPUT), fill=FILL_YELLOW, border=True)
    dv_pk = DataValidation(type="list", formula1="'Ref_Calendar'!$A$2:$A$500",
                           allow_blank=True)
    ws.add_data_validation(dv_pk)
    dv_pk.add("C7")
    dv_maj = DataValidation(type="list", formula1=f'"{MAJCOMS}"', allow_blank=True)
    ws.add_data_validation(dv_maj)
    dv_maj.add("C5")

    cell(ws, "A9", "HEADLINE — WHAT'S BROKEN", f=font(bold=True, size=12),
         fill=FILL_HDR)
    cell(ws, "A10", "DFACs out of tolerance:", f=font(bold=True))
    cell(ws, "D10", f'=SUMPRODUCT(({rP}="FAIL")*({rY}=1))', f=font(bold=True),
         fmt=FMT_INT)
    cell(ws, "A11", "Exceptions open:", f=font())
    cell(ws, "D11", '="no data (Stage 3)"', f=font(italic=True, color="FF808080"))
    cell(ws, "A12", "Dollars at risk:", f=font())
    cell(ws, "D12", '="no data (Stage 3)"', f=font(italic=True, color="FF808080"))

    cell(ws, "A14", "ESCALATION STATUS", f=font(bold=True, size=12), fill=FILL_HDR)
    cell(ws, "A15", "Month 1 — MFR due (manager):", f=font())
    cell(ws, "D15", f'=SUMPRODUCT(({rO}=1)*({rY}=1))', fmt=FMT_INT)
    cell(ws, "A16", "Month 2 — FSS/CC signature required:", f=font())
    cell(ws, "D16", f'=SUMPRODUCT(({rO}=2)*({rY}=1))', fmt=FMT_INT)
    cell(ws, "A17", "Month 3+ — investigation triggered:", f=font())
    cell(ws, "D17", f'=SUMPRODUCT(({rO}>=3)*({rY}=1))', fmt=FMT_INT)

    # helper counters
    cell(ws, "H19", f'=SUMPRODUCT(({rS}<>"")*({rY}=1))', f=font(color="FF808080"),
         fmt=FMT_INT)   # watch count
    cell(ws, "G19", "watch count ►", f=font(color="FF808080", size=8),
         align="right")

    cell(ws, "A19", "WATCH LIST — one more miss triggers investigation",
         f=font(bold=True, size=12), fill=FILL_HDR)
    for i, (col, h) in enumerate([("A", "DFAC_ID"), ("B", "DFAC_NAME"),
                                  ("C", "BASE_ID"), ("D", "MAJCOM"),
                                  ("E", "ESCALATION")]):
        cell(ws, f"{col}20", h, f=font(bold=True, color="FFFFFFFF"),
             fill=FILL_TITLE, border=True)
    for k in range(1, 13):
        r = 20 + k
        if k == 1:
            aformula = (f'=IF($H$19=0,"Nothing on the watch list — you\'re clear.",'
                        f'IFERROR(INDEX({rA},MATCH(1,{rZ},0)),""))')
        else:
            aformula = f'=IFERROR(INDEX({rA},MATCH({k},{rZ},0)),"")'
        cell(ws, f"A{r}", aformula, border=True)
        cell(ws, f"B{r}", f'=IFERROR(INDEX(\'Ref_Crosswalk\'!$B$2:$B$500,'
             f'MATCH(INDEX({rA},MATCH({k},{rZ},0)),'
             f'\'Ref_Crosswalk\'!$A$2:$A$500,0)),"")', border=True)
        cell(ws, f"C{r}", f'=IFERROR(INDEX({CT}$X$8:$X$307,MATCH({k},{rZ},0)),"")',
             border=True)
        cell(ws, f"D{r}", f'=IFERROR(INDEX({CT}$W$8:$W$307,MATCH({k},{rZ},0)),"")',
             border=True)
        cell(ws, f"E{r}", f'=IFERROR(INDEX({CT}$R$8:$R$307,MATCH({k},{rZ},0)),"")',
             border=True)

    cell(ws, "A34", "DATA INTEGRITY", f=font(bold=True, size=12), fill=FILL_HDR)
    cell(ws, "A35", "Provenance completeness:", f=font())
    cell(ws, "D35", '="no data (Stage 2)"', f=font(italic=True, color="FF808080"))
    cell(ws, "A36", "Traceability:", f=font())
    cell(ws, "D36", '="no data (Stage 2)"', f=font(italic=True, color="FF808080"))
    cell(ws, "A37", "Sources ingested:", f=font())
    cell(ws, "D37", '="no data (Stage 2)"', f=font(italic=True, color="FF808080"))

    cell(ws, "H39", f'=SUMPRODUCT(({rT}<>"")*({rY}=1))', f=font(color="FF808080"),
         fmt=FMT_INT)   # overdue count
    cell(ws, "G39", "overdue count ►", f=font(color="FF808080", size=8),
         align="right")
    cell(ws, "A39", "ACTIONS OVERDUE", f=font(bold=True, size=12), fill=FILL_HDR)
    for col, h in [("A", "DFAC_ID"), ("B", "PERIOD_KEY"), ("C", "WHAT'S OVERDUE")]:
        cell(ws, f"{col}40", h, f=font(bold=True, color="FFFFFFFF"),
             fill=FILL_TITLE, border=True)
    for k in range(1, 13):
        r = 40 + k
        if k == 1:
            aformula = (f'=IF($H$39=0,"No overdue actions.",'
                        f'IFERROR(INDEX({rA},MATCH(1,{rAA},0)),""))')
        else:
            aformula = f'=IFERROR(INDEX({rA},MATCH({k},{rAA},0)),"")'
        cell(ws, f"A{r}", aformula, border=True)
        cell(ws, f"B{r}", f'=IFERROR(INDEX({CT}$B$8:$B$307,MATCH({k},{rAA},0)),"")',
             border=True)
        cell(ws, f"C{r}", f'=IFERROR(INDEX({CT}$T$8:$T$307,MATCH({k},{rAA},0)),"")',
             border=True)

    cell(ws, "A54", "Color means one thing: RED = act now, AMBER = watch. "
         "One button: Data > Refresh All. Filters recompute every count.",
         f=font(italic=True, size=9))

    # conditional formatting on headline
    red = PatternFill("solid", fgColor=FILL_RED)
    amber = PatternFill("solid", fgColor=FILL_AMBER)
    ws.conditional_formatting.add("D10",
        CellIsRule(operator="greaterThan", formula=["0"], fill=red))
    ws.conditional_formatting.add("D15",
        CellIsRule(operator="greaterThan", formula=["0"], fill=amber))
    ws.conditional_formatting.add("D16",
        CellIsRule(operator="greaterThan", formula=["0"], fill=amber))
    ws.conditional_formatting.add("D17",
        CellIsRule(operator="greaterThan", formula=["0"], fill=red))

    widths = {"A": 26, "B": 22, "C": 16, "D": 16, "E": 34, "F": 4, "G": 16, "H": 8}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w
    ws.freeze_panes = "A3"


# -------------------------------------------------------------------------
# 9. Out_DataDictionary (§8.5)
# -------------------------------------------------------------------------
DD_ROWS = [
    # TAB, COLUMN, TYPE, SOURCE, DEFINITION, VALIDATION, OWNER
    ("Ref_Tolerance", "FACILITY_TYPE", "Text", "Seed", "Join key to Ref_Crosswalk", "Unique", "AFSVC/VMF"),
    ("Ref_Tolerance", "TOLERANCE_PCT", "Number", "DAFMAN 34-131", "Gain/loss tolerance as a fraction (0.03 = 3%)", "Regulatory — do not edit", "AFSVC/VMF"),
    ("Ref_Tolerance", "BASIS", "Text", "DAFMAN 34-131", "Income basis the tolerance applies against", "Enum", "AFSVC/VMF"),
    ("Ref_Tolerance", "METRIC_TYPE", "Text", "DAFMAN 34-131", "VARIANCE or MARGIN; Food 2.0 is MARGIN", "VARIANCE/MARGIN", "AFSVC/VMF"),
    ("Ref_Tolerance", "DAFMAN_CITE", "Text", "DAFMAN 34-131", "Paragraph the standard comes from", "-", "AFSVC/VMF"),
    ("Ref_Tolerance", "TIER", "Text", "DAFMAN 34-131", "Waiver tier (e.g. T-2)", "-", "AFSVC/VMF"),
    ("Ref_Tolerance", "NOTES", "Text", "Analyst", "Surfaces the 3% vs 5% conflict and the Food 2.0 gap", "-", "AFSVC/VMF"),
    ("Ref_Calendar", "PERIOD_KEY", "Text", "Derived", "FY{n}-P{MM}{A|B}; A=1-15, B=16-EOM", "Unique", "AFSVC/VMF"),
    ("Ref_Calendar", "FY", "Number", "Derived", "Federal fiscal year (Oct-Sep)", "-", "AFSVC/VMF"),
    ("Ref_Calendar", "MONTH", "Number", "Derived", "Calendar month number", "1-12", "AFSVC/VMF"),
    ("Ref_Calendar", "HALF", "Text", "Derived", "A (1-15) or B (16-EOM)", "A/B", "AFSVC/VMF"),
    ("Ref_Calendar", "START_DATE/END_DATE", "Date", "Derived", "Period bounds", "-", "AFSVC/VMF"),
    ("Ref_Calendar", "INVENTORY_DUE", "Date", "para 7.13", "15th or last day of month", "-", "AFSVC/VMF"),
    ("Ref_Calendar", "POST_DUE", "Date", "para 7.13", "INVENTORY_DUE + 5 days", "-", "AFSVC/VMF"),
    ("Ref_Calendar", "MMR_DUE", "Date", "para 5.2.5", "10th of following month", "-", "AFSVC/VMF"),
    ("Ref_Calendar", "STATUS", "Text", "Maintained", "OPEN or CLOSED", "OPEN/CLOSED", "Build owner"),
    ("Ref_Crosswalk", "DFAC_ID", "Text", "AFSVC/VMF", "Canonical primary key {BASE_ID}-{NN}", "Unique, immutable", "AFSVC/VMF"),
    ("Ref_Crosswalk", "DFAC_NAME/BASE_*/MAJCOM", "Text", "AFSVC/VMF", "Facility identity fields", "MAJCOM from list", "AFSVC/VMF"),
    ("Ref_Crosswalk", "FACILITY_TYPE", "Text", "AFSVC/VMF", "Drives tolerance selection", "In Ref_Tolerance", "AFSVC/VMF"),
    ("Ref_Crosswalk", "DODAAC", "Text", "para 1.4.10", "DoD Activity Address Code", "-", "AFSVC/VMF"),
    ("Ref_Crosswalk", "ALOHA_STORE/CT_SITE/CIR_MERCHANT/STORES_ACCT", "Text", "Pending", "Source-system identifiers (placeholders)", "Yellow — pending", "AFSVC/VMF"),
    ("Ref_Crosswalk", "ACTIVE/EFF_START/EFF_END", "Y-N/Date", "AFSVC/VMF", "Effective-dating for renames/renumbering", "-", "AFSVC/VMF"),
    ("Ref_ExceptionTypes", "EXC_CODE..REMEDIATION", "Text", "Seed §5.6", "Exception taxonomy and routing", "-", "AFSVC/VMF"),
    ("Ref_AdapterManifest", "QUERY_NAME..NOTES", "Mixed", "Build owner", "Transformation edge list; adapters are Stage 2+", "STATUS enum", "Build owner"),
    ("Ref_CrosswalkVersion", "CROSSWALK_VER..ROW_COUNT", "Mixed", "AFSVC/VMF", "Crosswalk change history for drift detection", "-", "AFSVC/VMF"),
    ("Out_ComplianceTracker", "DFAC_ID", "Text", "Input", "Dropdown from Ref_Crosswalk", "In crosswalk", "Accountant"),
    ("Out_ComplianceTracker", "PERIOD_KEY", "Text", "Input", "Dropdown from Ref_Calendar", "In calendar", "Accountant"),
    ("Out_ComplianceTracker", "EARNED_INCOME", "Number", "MMR", "Monthly earned income", ">=0", "Accountant"),
    ("Out_ComplianceTracker", "GAIN_LOSS_AMT", "Number", "MMR", "Signed gain/loss amount", "-", "Accountant"),
    ("Out_ComplianceTracker", "COST_OF_GOODS", "Number", "MMR", "Food 2.0 only", ">=0", "Accountant"),
    ("Out_ComplianceTracker", "MFR/FSS_CC/INVESTIGATION", "Y/N", "Input", "Escalation actions taken", "Y/N", "Accountant"),
    ("Out_ComplianceTracker", "FACILITY_TYPE", "Formula", "Ref_Crosswalk", "Looked up from crosswalk", "-", "System"),
    ("Out_ComplianceTracker", "TOLERANCE_PCT/METRIC_TYPE/DAFMAN_CITE", "Formula", "Ref_Tolerance", "Applied standard and its cite", "-", "System"),
    ("Out_ComplianceTracker", "VAR_PCT", "Formula", "Computed", "ABS(gain/loss)/earned income, guarded", "-", "System"),
    ("Out_ComplianceTracker", "COG_PCT", "Formula", "Computed", "Cost of goods / earned income (Food 2.0)", "-", "System"),
    ("Out_ComplianceTracker", "STREAK", "Formula", "Computed", "Consecutive FAIL periods, resets on PASS", "-", "System"),
    ("Out_ComplianceTracker", "RESULT", "Formula", "Computed", "PASS/FAIL/CHECK; MARGIN vs VARIANCE branch", "-", "System"),
    ("Out_ComplianceTracker", "ESCALATION/WATCH/ACTION_OVERDUE", "Formula", "Computed", "Ladder state and preventive flags (para 5.13.2)", "-", "System"),
    ("Out_ComplianceTracker", "SEQ..OVERDUE_RANK", "Formula", "Helper", "Internal helpers for streak and dashboard lists", "-", "System"),
    ("Out_Dashboard", "Filters C5:C7", "Input", "User", "MAJCOM / BASE_ID / PERIOD_KEY filters (blank=all)", "-", "Any"),
    ("Out_Dashboard", "Counts & lists", "Formula", "Out_ComplianceTracker", "Escalation counts, watch list, overdue list", "-", "System"),
    ("Out_ValidationRegister", "Rule results", "Formula", "Workbook", "PASS/FAIL/N-A per validation rule", "-", "System"),
]


def build_out_datadictionary():
    ws = wb.create_sheet("Out_DataDictionary")
    ws.sheet_properties.tabColor = TAB_DARKBLUE
    title(ws, "DATA DICTIONARY", "Every column: type, source, definition, validation, owner.")
    heads = ["TAB", "COLUMN", "TYPE", "SOURCE", "DEFINITION", "VALIDATION", "OWNER"]
    header_row(ws, 3, heads)
    for i, row in enumerate(DD_ROWS):
        r = 4 + i
        for j, val in enumerate(row):
            col = get_column_letter(1 + j)
            cell(ws, f"{col}{r}", val, fill=(FILL_SEED if i % 2 else None),
                 border=True, wrap=(j in (4,)))
    widths = {"A": 22, "B": 34, "C": 10, "D": 16, "E": 46, "F": 18, "G": 13}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w
    ws.freeze_panes = "A4"


# -------------------------------------------------------------------------
# 10. Out_ValidationRegister (§8.6) — 14 numbered rules
# -------------------------------------------------------------------------
def build_out_validationregister():
    ws = wb.create_sheet("Out_ValidationRegister")
    ws.sheet_properties.tabColor = TAB_DARKBLUE
    title(ws, "VALIDATION REGISTER",
          "Numbered rules with live PASS/FAIL. Ledger/provenance rules activate in "
          "Stage 2+.")
    heads = ["#", "RULE", "SCOPE", "RESULT", "DETAIL"]
    header_row(ws, 3, heads)
    CT = "'Out_ComplianceTracker'!"
    TOL = "'Ref_Tolerance'!"
    CAL = "'Ref_Calendar'!"
    CW = "'Ref_Crosswalk'!"

    live = {}
    live[1] = (f'=IF(SUMPRODUCT(({CT}$A$8:$A$307<>"")*'
               f'({CT}$J$8:$J$307="UNKNOWN"))=0,"PASS","FAIL")')
    live[4] = (f'=IF(SUMPRODUCT(({CT}$B$8:$B$307<>"")*'
               f'(COUNTIF({CAL}$A$2:$A$500,{CT}$B$8:$B$307)=0))=0,"PASS","FAIL")')
    live[6] = (f'=IF(SUMPRODUCT(({CW}$F$2:$F$500<>"")*'
               f'(COUNTIF({TOL}$A$2:$A$9,{CW}$F$2:$F$500)=0))=0,"PASS","FAIL")')
    live[7] = (f'=IF(SUMPRODUCT(({TOL}$D$2:$D$9="VARIANCE")*'
               f'(ISNUMBER({TOL}$B$2:$B$9)=FALSE))=0,"PASS","FAIL")')
    live[8] = (f'=IF(SUMPRODUCT(({CT}$P$8:$P$307="PASS")*'
               f'({CT}$O$8:$O$307<>0))=0,"PASS","FAIL")')
    live[9] = (f'=IF(SUMPRODUCT(({CT}$A$8:$A$307<>"")*('
               f'(({CT}$O$8:$O$307=0)*({CT}$R$8:$R$307<>"CLEAR"))+'
               f'(({CT}$O$8:$O$307=1)*({CT}$R$8:$R$307<>"MFR DUE - MANAGER"))+'
               f'(({CT}$O$8:$O$307=2)*({CT}$R$8:$R$307<>'
               f'"MFR DUE - FSS/CC SIGNATURE REQUIRED"))+'
               f'(({CT}$O$8:$O$307>=3)*({CT}$R$8:$R$307<>'
               f'"INVESTIGATION + REPORT OF SURVEY; MSG/CC NOTIFIED"))))=0,'
               f'"PASS","FAIL")')
    live[10] = (f'=IF(SUMPRODUCT(--ISERROR({CT}$J$8:$AA$307))+'
                f'SUMPRODUCT(--ISERROR(\'Out_Dashboard\'!$A$10:$E$52))=0,'
                f'"PASS","FAIL")')
    live[11] = (f'=IF(SUMPRODUCT(({CAL}$A$2:$A$500<>"")*('
                f'(({CAL}$D$2:$D$500="A")*(DAY({CAL}$G$2:$G$500)<>15))+'
                f'(({CAL}$D$2:$D$500="B")*({CAL}$G$2:$G$500<>{CAL}$F$2:$F$500))))=0,'
                f'"PASS","FAIL")')

    rules = [
        (1, "All DFAC_ID in the tracker resolve in Ref_Crosswalk", "LIVE",
         "No FACILITY_TYPE lookups return UNKNOWN."),
        (2, "Provenance completeness = 100%", "STAGE 2+",
         "Activates with Ledger_ProvenanceCheck."),
        (3, "All ADAPTER_VER present in Ref_AdapterManifest as ACTIVE", "STAGE 2+",
         "Activates with adapters."),
        (4, "No PERIOD_KEY outside Ref_Calendar", "LIVE",
         "Every tracker PERIOD_KEY exists in the calendar."),
        (5, "Sum of adapter row counts = ledger row count", "STAGE 2+",
         "Activates with the ledger."),
        (6, "Every FACILITY_TYPE resolves in Ref_Tolerance", "LIVE",
         "Catches misclassified/unknown facility types (TC-09)."),
        (7, "No TOLERANCE_PCT blank where METRIC_TYPE = VARIANCE", "LIVE",
         "Every variance rule carries a numeric tolerance."),
        (8, "Streak counter resets on PASS", "LIVE",
         "STREAK = 0 on every PASS row (TC-06)."),
        (9, "Escalation state consistent with streak value", "LIVE",
         "Recomputed ladder text matches ESCALATION."),
        (10, "No naked formula errors anywhere", "LIVE",
         "No error cells in tracker calc range or dashboard (TC-22)."),
        (11, "Inventory count dates align to 15th / EOM", "LIVE",
         "Calendar INVENTORY_DUE matches para 7.13."),
        (12, "AMOUNT = QTY x UNIT_COST within tolerance", "STAGE 2+",
         "Activates with the ledger."),
        (13, "No duplicate TXN_ID", "STAGE 2+",
         "Activates with the ledger."),
        (14, "All EXC_CODE resolve in Ref_ExceptionTypes", "STAGE 2+",
         "Activates with the reconciliation layer."),
    ]
    for i, (num, text, scope, detail) in enumerate(rules):
        r = 4 + i
        cell(ws, f"A{r}", num, border=True, align="center")
        cell(ws, f"B{r}", text, border=True, wrap=True)
        cell(ws, f"C{r}", scope, border=True, align="center",
             f=font(color=("FF808080" if scope != "LIVE" else C_FORMULA)))
        if num in live:
            cell(ws, f"D{r}", live[num], border=True, align="center",
                 f=font(bold=True))
        else:
            cell(ws, f"D{r}", "N/A — Stage 2+", border=True, align="center",
                 f=font(italic=True, color="FF808080"))
        cell(ws, f"E{r}", detail, border=True, wrap=True, f=font(size=9))
    # summary
    sr = 4 + len(rules) + 1
    cell(ws, f"B{sr}", "Live rules passing:", f=font(bold=True))
    cell(ws, f"D{sr}", '="  "&SUMPRODUCT(--($D$4:$D$17="PASS"))&" of "'
         '&SUMPRODUCT(--($C$4:$C$17="LIVE"))', f=font(bold=True))
    # conditional formatting: FAIL red, PASS green tint
    red = PatternFill("solid", fgColor=FILL_RED)
    green = PatternFill("solid", fgColor="FFC6EFCE")
    ws.conditional_formatting.add("D4:D17",
        CellIsRule(operator="equal", formula=['"FAIL"'], fill=red))
    ws.conditional_formatting.add("D4:D17",
        CellIsRule(operator="equal", formula=['"PASS"'], fill=green))
    widths = {"A": 5, "B": 46, "C": 11, "D": 16, "E": 44}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w
    ws.freeze_panes = "A4"


# ---- Build in layer/tab order (spec §9.1) --------------------------------
build_ref_tolerance()
build_ref_calendar()
build_ref_crosswalk()
build_ref_exceptiontypes()
build_ref_adaptermanifest()
build_ref_crosswalkversion()
build_out_compliancetracker()
build_out_dashboard()
build_out_datadictionary()
build_out_validationregister()

# Force full recalculation on open so Excel/LibreOffice compute (no cached values)
try:
    wb.calculation.fullCalcOnLoad = True
except Exception:
    pass

out = sys.argv[1] if len(sys.argv) > 1 else "MFRP_v1.xlsx"
wb.save(out)
print("Saved", out)
